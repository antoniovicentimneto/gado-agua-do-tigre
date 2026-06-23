"""Testes da restauração a partir do Excel exportado pelo app."""
import io

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.models import Animal
from app.services.exportacao import gerar_planilha
from app.services.restauracao import restaurar_backup


@pytest.fixture
def db_vazio():
    """Banco SQLite em memória sem nenhum animal (simula um app recém-instalado)."""
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    s = sessionmaker(bind=engine)()
    yield s
    s.close()


def test_restaura_tudo_num_banco_vazio(db, db_vazio, tmp_path):
    conteudo = gerar_planilha(db)  # `db` (conftest) tem 3 animais semeados
    caminho = tmp_path / "backup.xlsx"
    caminho.write_bytes(conteudo)

    resumo = restaurar_backup(str(caminho), db_vazio)
    assert resumo["animais_novos"] == 3
    assert resumo["pesagens_novas"] == 4  # 2 + 1 + 1 pesagens semeadas no fixture

    brincos = {a.brinco for a in db_vazio.query(Animal).all()}
    assert brincos == {"101", "102", "201"}


def test_brincos_duplicados_continuam_animais_distintos(db_vazio, tmp_path):
    # Monta um backup à mão com 2 animais diferentes e o MESMO brinco.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Animais"
    ws.append(["Brinco", "Tipo", "Raça", "Cor", "Lote atual", "Situação", "Capado",
               "Nascimento", "Observação"])
    ws.append(["50", "Boi", None, None, "LOTEA", "ativo", "", None, None])
    ws.append(["50", "Vaca", None, None, "LOTEB", "ativo", "", None, None])
    ws2 = wb.create_sheet("Pesos")
    ws2.append(["Brinco", "01/01/2026"])
    ws2.append(["50", 300])
    ws2.append(["50", 450])
    caminho = tmp_path / "backup_dup.xlsx"
    wb.save(caminho)

    resumo = restaurar_backup(str(caminho), db_vazio)
    assert resumo["animais_novos"] == 2
    assert resumo["pesagens_novas"] == 2

    animais = db_vazio.query(Animal).filter(Animal.brinco == "50").all()
    assert len(animais) == 2
    assert {a.tipo for a in animais} == {"Boi", "Vaca"}
    assert {a.pesagens[0].peso for a in animais} == {300.0, 450.0}


def test_recusa_restaurar_em_banco_que_ja_tem_animal(db, tmp_path):
    conteudo = gerar_planilha(db)
    caminho = tmp_path / "backup.xlsx"
    caminho.write_bytes(conteudo)

    with pytest.raises(ValueError, match="já tem animais"):
        restaurar_backup(str(caminho), db)


def test_recusa_arquivo_sem_as_abas_certas(db_vazio, tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "OutraCoisa"
    caminho = tmp_path / "errado.xlsx"
    wb.save(caminho)

    with pytest.raises(ValueError, match="Animais"):
        restaurar_backup(str(caminho), db_vazio)

"""Testes da exportação Excel e da importação incremental."""
import io

import openpyxl

from app.models import Animal, Pesagem
from app.services.exportacao import gerar_planilha


def test_exportar_excel_tem_abas_e_dados(db):
    conteudo = gerar_planilha(db)
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    assert wb.sheetnames == ["Animais", "Pesos"]
    # 3 animais semeados + cabeçalho.
    assert wb["Animais"].max_row == 4
    # Aba Pesos tem a coluna Brinco + as datas.
    assert wb["Pesos"].cell(1, 1).value == "Brinco"
    assert wb["Pesos"].max_column >= 2


def test_incremental_nao_apaga_e_so_adiciona(db, tmp_path):
    # Lança um peso "do app" numa data que a planilha não terá.
    from datetime import date
    a101 = db.query(Animal).filter(Animal.brinco == "101").first()
    db.add(Pesagem(animal_id=a101.id, data=date(2026, 6, 20), peso=999))
    db.commit()

    # Monta uma planilha mínima no formato DADOS com um animal novo e uma data nova.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DADOS"
    # Cabeçalho: col G(7)=Brinco, col X(24)=primeira data.
    ws.cell(1, 7, "Brinco")
    ws.cell(1, 11, "Tipo")
    ws.cell(1, 24, "10/06/2026")
    # Linha de um animal NOVO (brinco 555) com peso na data nova.
    ws.cell(2, 7, "555")
    ws.cell(2, 11, "Boi")
    ws.cell(2, 24, 250)
    caminho = tmp_path / "mini.xlsx"
    wb.save(caminho)

    from app.services.importacao import importar_incremental
    r = importar_incremental(str(caminho), db)
    assert r["animais_novos"] == 1          # criou o 555
    # O peso "do app" (999 em 2026-06-20) continua lá.
    db.refresh(a101)
    assert any(p.peso == 999 for p in a101.pesagens)

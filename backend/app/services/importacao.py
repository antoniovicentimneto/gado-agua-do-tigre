"""Importa a planilha antiga (aba DADOS) para o banco do app.

A planilha guarda cada data de pesagem como uma COLUNA. Aqui essas colunas viram
linhas na tabela `pesagens` (uma linha por pesagem). Também corrige datas digitadas
com erro e mapeia status, lote, dentição e score.
"""
from __future__ import annotations

import re
from datetime import date, datetime

import openpyxl
from sqlalchemy.orm import Session

from ..models import (
    Animal,
    AnimalLote,
    Compra,
    Denticao,
    Lote,
    Pesagem,
    Score,
    StatusAnimal,
)

# Mapeamento das colunas de cadastro (1-based) conforme a aba DADOS.
COL_OBSERVACAO = 1   # A
COL_NASCIMENTO = 3   # C  (DN BEZ)
COL_DENTICAO = 4     # D
COL_SCORE = 5        # E
COL_BRINCO = 7       # G
COL_SITUACAO = 9     # I
COL_LOTE = 10        # J
COL_TIPO = 11        # K
COL_VENDEDOR = 12    # L
COL_RACA = 13        # M
COL_COR = 14         # N
COL_VALOR = 17       # Q
COL_KG_COMPRA = 21   # U

# A partir de qual coluna começam as datas de pesagem (X = 24).
PRIMEIRA_COL_DATA = 24


def _texto(valor) -> str | None:
    """Normaliza um valor de célula em texto limpo (ou None)."""
    if valor is None:
        return None
    s = str(valor).strip()
    return s or None


def _parse_data(valor) -> date | None:
    """Interpreta o cabeçalho de uma coluna como data, corrigindo erros comuns.

    Retorna None se não parecer uma data (ex.: 'ultimo', 'FINAL', 'GMD').
    """
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor

    s = str(valor).strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,5})$", s)
    if not m:
        return None
    dia, mes, ano = (int(g) for g in m.groups())

    # Corrige ano digitado errado: '20232' -> 2023 (mantém os 4 primeiros dígitos).
    if ano > 9999:
        ano = int(str(ano)[:4])
    elif ano < 100:  # ano com 2 dígitos
        ano += 2000

    try:
        return date(ano, mes, dia)
    except ValueError:
        return None


def _parse_numero(valor) -> float | None:
    """Converte o valor de uma célula em número (kg, valor), tolerando texto."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _status_de(situacao: str | None) -> tuple[StatusAnimal, bool]:
    """Deriva (status, capado) a partir do texto livre da coluna Situação."""
    if not situacao:
        return StatusAnimal.ATIVO, False
    s = situacao.strip().upper()
    capado = "CAPADO" in s
    if "VENDID" in s:
        return StatusAnimal.VENDIDO, capado
    if "MORR" in s or "MORTO" in s:
        return StatusAnimal.MORTO, capado
    return StatusAnimal.ATIVO, capado


def importar_planilha(caminho: str, db: Session) -> dict:
    """Lê a planilha e popula o banco. Retorna um resumo com contadores."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb["DADOS"]

    # 1) Mapeia as colunas-data a partir do cabeçalho (linha 1).
    colunas_data: dict[int, date] = {}
    datas_corrigidas: list[str] = []
    for col in range(PRIMEIRA_COL_DATA, ws.max_column + 1):
        cabecalho = ws.cell(1, col).value
        d = _parse_data(cabecalho)
        if d is None:
            continue
        # Guarda aviso quando o texto original não bate com a data interpretada.
        if isinstance(cabecalho, str) and not re.match(
            r"^\d{1,2}/\d{1,2}/\d{4}$", cabecalho.strip()
        ):
            datas_corrigidas.append(f"{cabecalho!r} -> {d.isoformat()}")
        colunas_data[col] = d

    # Cache de lotes para não duplicar.
    lotes_cache: dict[str, Lote] = {}

    def obter_lote(nome: str) -> Lote:
        lote = lotes_cache.get(nome)
        if lote is None:
            lote = Lote(nome=nome)
            db.add(lote)
            db.flush()
            lotes_cache[nome] = lote
        return lote

    resumo = {
        "animais": 0,
        "pesagens": 0,
        "denticoes": 0,
        "scores": 0,
        "lotes": 0,
        "datas_corrigidas": datas_corrigidas,
        "pulados_sem_brinco": 0,
        "pesagens_duplicadas": 0,
    }

    for r in range(2, ws.max_row + 1):
        brinco = _texto(ws.cell(r, COL_BRINCO).value)
        if not brinco:
            resumo["pulados_sem_brinco"] += 1
            continue

        situacao = _texto(ws.cell(r, COL_SITUACAO).value)
        status, capado = _status_de(situacao)

        # Observação: junta a observação livre + situação original quando não é status.
        obs_partes = []
        obs = _texto(ws.cell(r, COL_OBSERVACAO).value)
        if obs:
            obs_partes.append(obs)
        if situacao and status == StatusAnimal.ATIVO and not capado:
            obs_partes.append(f"(planilha: {situacao})")

        animal = Animal(
            brinco=brinco,
            tipo=_texto(ws.cell(r, COL_TIPO).value),
            raca=_texto(ws.cell(r, COL_RACA).value),
            cor=_texto(ws.cell(r, COL_COR).value),
            vendedor=_texto(ws.cell(r, COL_VENDEDOR).value),
            nascimento=_parse_data(ws.cell(r, COL_NASCIMENTO).value),
            capado=capado,
            status=status,
            observacao=" ".join(obs_partes) or None,
        )
        db.add(animal)
        db.flush()  # garante animal.id
        resumo["animais"] += 1

        # Pesagens: percorre as colunas-data; dedup por data (mantém o último valor).
        pesos_por_data: dict[date, float] = {}
        for col, d in colunas_data.items():
            peso = _parse_numero(ws.cell(r, col).value)
            if peso is None or peso <= 0:
                continue
            if d in pesos_por_data:
                resumo["pesagens_duplicadas"] += 1
            pesos_por_data[d] = peso

        for d, peso in sorted(pesos_por_data.items()):
            db.add(Pesagem(animal_id=animal.id, data=d, peso=peso))
            resumo["pesagens"] += 1

        ultima_data = max(pesos_por_data) if pesos_por_data else None

        # Dentição (planilha tem só um número, sem data): usa a última pesagem como referência.
        dentes = _parse_numero(ws.cell(r, COL_DENTICAO).value)
        if dentes is not None and ultima_data is not None:
            db.add(Denticao(animal_id=animal.id, data=ultima_data, dentes=int(dentes)))
            resumo["denticoes"] += 1

        # Score corporal (idem: usa a última pesagem como data de referência).
        score = _parse_numero(ws.cell(r, COL_SCORE).value)
        if score is not None and ultima_data is not None:
            db.add(Score(animal_id=animal.id, data=ultima_data, valor=score))
            resumo["scores"] += 1

        # Lote atual (histórico ainda sem datas — começa com o lote da planilha).
        nome_lote = _texto(ws.cell(r, COL_LOTE).value)
        if nome_lote:
            lote = obter_lote(nome_lote)
            db.add(AnimalLote(animal_id=animal.id, lote_id=lote.id))

        # Compra (data desconhecida na planilha).
        kg = _parse_numero(ws.cell(r, COL_KG_COMPRA).value)
        valor = _parse_numero(ws.cell(r, COL_VALOR).value)
        if kg is not None or valor is not None:
            db.add(Compra(animal_id=animal.id, kg=kg, valor=valor))

    db.commit()
    resumo["lotes"] = len(lotes_cache)
    return resumo


def _mapear_colunas_data(ws) -> dict[int, date]:
    """Detecta as colunas-data do cabeçalho (reaproveitado na importação incremental)."""
    colunas: dict[int, date] = {}
    for col in range(PRIMEIRA_COL_DATA, ws.max_column + 1):
        d = _parse_data(ws.cell(1, col).value)
        if d is not None:
            colunas[col] = d
    return colunas


def importar_incremental(caminho: str, db: Session) -> dict:
    """Atualiza o banco a partir da planilha SEM apagar nada.

    - Animais novos (brincos que não existem) são criados.
    - Pesagens em datas que o animal ainda não tem são adicionadas.
    - Nada é removido e o lote/dados de animais já existentes NÃO são alterados
      (protege o que foi lançado/movido dentro do app).
    """
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb["DADOS"]
    colunas_data = _mapear_colunas_data(ws)

    def obter_ou_criar_lote(nome: str) -> Lote:
        lote = db.query(Lote).filter(Lote.nome == nome).first()
        if lote is None:
            lote = Lote(nome=nome)
            db.add(lote)
            db.flush()
        return lote

    resumo = {"animais_novos": 0, "pesagens_novas": 0, "animais_existentes": 0,
              "pulados_sem_brinco": 0}

    for r in range(2, ws.max_row + 1):
        brinco = _texto(ws.cell(r, COL_BRINCO).value)
        if not brinco:
            resumo["pulados_sem_brinco"] += 1
            continue

        animal = db.query(Animal).filter(Animal.brinco == brinco).first()
        novo = animal is None
        if novo:
            situacao = _texto(ws.cell(r, COL_SITUACAO).value)
            status, capado = _status_de(situacao)
            animal = Animal(
                brinco=brinco,
                tipo=_texto(ws.cell(r, COL_TIPO).value),
                raca=_texto(ws.cell(r, COL_RACA).value),
                cor=_texto(ws.cell(r, COL_COR).value),
                vendedor=_texto(ws.cell(r, COL_VENDEDOR).value),
                nascimento=_parse_data(ws.cell(r, COL_NASCIMENTO).value),
                capado=capado,
                status=status,
                observacao=_texto(ws.cell(r, COL_OBSERVACAO).value),
            )
            db.add(animal)
            db.flush()
            resumo["animais_novos"] += 1
            nome_lote = _texto(ws.cell(r, COL_LOTE).value)
            if nome_lote:
                db.add(AnimalLote(animal_id=animal.id, lote_id=obter_ou_criar_lote(nome_lote).id))
            kg = _parse_numero(ws.cell(r, COL_KG_COMPRA).value)
            valor = _parse_numero(ws.cell(r, COL_VALOR).value)
            if kg is not None or valor is not None:
                db.add(Compra(animal_id=animal.id, kg=kg, valor=valor))
        else:
            resumo["animais_existentes"] += 1

        # Adiciona apenas as pesagens de datas que o animal ainda não tem.
        datas_existentes = {p.data for p in animal.pesagens}
        pesos_por_data: dict[date, float] = {}
        for col, d in colunas_data.items():
            peso = _parse_numero(ws.cell(r, col).value)
            if peso is not None and peso > 0:
                pesos_por_data[d] = peso
        for d, peso in sorted(pesos_por_data.items()):
            if d not in datas_existentes:
                db.add(Pesagem(animal_id=animal.id, data=d, peso=peso))
                resumo["pesagens_novas"] += 1

    db.commit()
    return resumo

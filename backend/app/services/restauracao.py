"""Restaura o banco a partir do Excel que o PRÓPRIO app exportou (Painel → Exportar).

Serve pra recuperar tudo (cadastro + lotes + todas as pesagens) se o app for
desinstalado/reinstalado ou o banco for perdido. Diferente do `atualizar.py`
(que lê a planilha original de origem, aba DADOS), este lê o formato gerado
por `exportacao.gerar_planilha` (abas "Animais" e "Pesos").

Importante: o brinco NÃO é único na base (existem animais diferentes com o
mesmo número de brinco). As duas abas são geradas na mesma ordem, uma linha
por animal — por isso a linha N da aba "Animais" e a linha N da aba "Pesos"
são sempre o MESMO animal, e é assim que essa restauração faz o pareamento
(nunca procurando por brinco, que pode ser ambíguo).

Por segurança, só funciona contra um banco SEM nenhum animal cadastrado
(restauração "do zero" depois de perder a base). Pra mesclar uma planilha
numa base que já tem dados, use o `atualizar.py`.
"""
from __future__ import annotations

from datetime import date, datetime

import openpyxl
from sqlalchemy.orm import Session

from ..models import Animal, AnimalLote, Lote, Pesagem, StatusAnimal


def _data(valor) -> date | None:
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    s = str(valor).strip()
    try:
        return datetime.strptime(s, "%d/%m/%Y").date()
    except ValueError:
        return None


def _status(valor) -> StatusAnimal:
    s = str(valor or "").strip().lower()
    for st in StatusAnimal:
        if st.value == s:
            return st
    return StatusAnimal.ATIVO


def restaurar_backup(caminho: str, db: Session) -> dict:
    """Lê o Excel exportado pelo app e recria o rebanho inteiro num banco vazio."""
    if db.query(Animal).first() is not None:
        raise ValueError(
            "O banco já tem animais cadastrados — a restauração só funciona "
            "num banco vazio (depois de reinstalar o app, por exemplo). "
            "Pra atualizar uma base que já existe, use o atualizar.py."
        )

    try:
        wb = openpyxl.load_workbook(caminho, data_only=True)
    except Exception as e:
        raise ValueError(f"Não consegui abrir o arquivo Excel: {e}") from e

    if "Animais" not in wb.sheetnames or "Pesos" not in wb.sheetnames:
        raise ValueError(
            "Esse arquivo não parece ser o backup exportado pelo app "
            '(faltam as abas "Animais" e/ou "Pesos").'
        )

    ws_animais = wb["Animais"]
    ws_pesos = wb["Pesos"]

    resumo = {"animais_novos": 0, "pesagens_novas": 0, "pulados_sem_brinco": 0}

    lotes_cache: dict[str, Lote] = {}

    def obter_ou_criar_lote(nome: str) -> Lote:
        lote = lotes_cache.get(nome)
        if lote is None:
            lote = Lote(nome=nome)
            db.add(lote)
            db.flush()
            lotes_cache[nome] = lote
        return lote

    # Linhas das duas abas, na mesma ordem (uma por animal).
    linhas_animais = list(ws_animais.iter_rows(min_row=2, values_only=True))
    linhas_pesos = list(ws_pesos.iter_rows(min_row=2, values_only=True))
    cabecalho_pesos = next(ws_pesos.iter_rows(min_row=1, max_row=1, values_only=True))
    datas_colunas = {i: _data(v) for i, v in enumerate(cabecalho_pesos) if i > 0 and _data(v)}

    for pos, linha in enumerate(linhas_animais):
        if not linha or not linha[0]:
            resumo["pulados_sem_brinco"] += 1
            continue
        (brinco, tipo, raca, cor, lote_atual, situacao, capado, nascimento,
         observacao, *_resto) = linha

        animal = Animal(
            brinco=str(brinco),
            tipo=tipo or None,
            raca=raca or None,
            cor=cor or None,
            nascimento=_data(nascimento),
            capado=str(capado or "").strip().lower() == "sim",
            status=_status(situacao),
            observacao=observacao or None,
        )
        db.add(animal)
        db.flush()
        resumo["animais_novos"] += 1

        if lote_atual:
            db.add(AnimalLote(animal_id=animal.id, lote_id=obter_ou_criar_lote(str(lote_atual)).id))

        # Pesagens dessa mesma linha (posição igual na aba Pesos).
        linha_pesos = linhas_pesos[pos] if pos < len(linhas_pesos) else None
        if not linha_pesos:
            continue
        for i, d in datas_colunas.items():
            peso = linha_pesos[i] if i < len(linha_pesos) else None
            if peso is None:
                continue
            try:
                peso = float(peso)
            except (TypeError, ValueError):
                continue
            if peso <= 0:
                continue
            db.add(Pesagem(animal_id=animal.id, data=d, peso=peso))
            resumo["pesagens_novas"] += 1

    db.commit()
    return resumo

"""Gera uma planilha .xlsx completa a partir do banco do app.

Serve como backup legível e como a "planilha de fácil acesso" do usuário,
sempre atualizada com os dados do app. Tem duas abas:
- "Animais": cadastro + indicadores (lote, status, GMD, último peso...).
- "Pesos": formato largo (uma coluna por data), igual à planilha original.
"""
from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from ..models import Animal
from .consultas import lote_atual, pontos_pesagem
from .gmd import resumo_animal

NEGRITO = Font(bold=True)


def gerar_planilha(db: Session) -> bytes:
    """Monta o arquivo .xlsx e devolve os bytes."""
    animais = db.query(Animal).order_by(Animal.brinco).all()
    wb = Workbook()

    # ----------------------------------------------------- Aba Animais
    ws = wb.active
    ws.title = "Animais"
    cabecalho = [
        "Brinco", "Tipo", "Raça", "Cor", "Lote atual", "Situação", "Capado",
        "Nascimento", "Observação", "Qtd pesagens", "Primeiro peso", "Último peso",
        "Data último", "GMD", "uGMD", "Compra kg", "Compra R$", "Venda R$",
    ]
    ws.append(cabecalho)
    for celula in ws[1]:
        celula.font = NEGRITO

    for a in animais:
        r = resumo_animal(pontos_pesagem(a))
        ws.append([
            a.brinco, a.tipo, a.raca, a.cor, lote_atual(a), a.status.value,
            "Sim" if a.capado else "", a.nascimento, a.observacao,
            r["qtde_pesagens"], r["primeiro_peso"], r["ultimo_peso"], r["data_ultimo"],
            r["gmd"], r["ugmd"],
            a.compra.kg if a.compra else None,
            a.compra.valor if a.compra else None,
            a.venda.valor_recebido if a.venda else None,
        ])

    # ----------------------------------------------------- Aba Pesos (formato largo)
    ws2 = wb.create_sheet("Pesos")
    datas = sorted({p.data for a in animais for p in a.pesagens})
    ws2.append(["Brinco"] + [d.strftime("%d/%m/%Y") for d in datas])
    for celula in ws2[1]:
        celula.font = NEGRITO
    idx = {d: i for i, d in enumerate(datas)}
    for a in animais:
        linha = [a.brinco] + [None] * len(datas)
        for p in a.pesagens:
            linha[idx[p.data] + 1] = p.peso
        ws2.append(linha)

    # Congela cabeçalho e primeira coluna nas duas abas.
    ws.freeze_panes = "A2"
    ws2.freeze_panes = "B2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def nome_arquivo() -> str:
    return f"gado_agua_do_tigre_{date.today().isoformat()}.xlsx"
